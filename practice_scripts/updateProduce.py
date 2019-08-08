#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl
from openpyxl.styles import Font

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

   # Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row):  # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
italic24Font = Font(size=24, italic=True)
calibri = Font(name='Calibri', size=33)
sheet.cell(row=1, column=1).font = calibri
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
sheet.freeze_panes = 'A2'
wb.save('updatedProduceSales.xlsx')